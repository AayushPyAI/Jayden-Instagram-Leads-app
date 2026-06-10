"""Gemini vision extraction + Excel deduplication and append."""

from __future__ import annotations

import asyncio
import io
import json
import os
import re
import threading
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import google.generativeai as genai
import pandas as pd
from PIL import Image
from openpyxl import load_workbook

from config import (
    BATCH_DEDUPE_SOURCE_LABEL,
    DEFAULT_GEMINI_MODEL,
    EXTRACTION_SYSTEM_INSTRUCTION,
    GEMINI_CONCURRENCY,
    GEMINI_DISCOVER_MODELS,
    GEMINI_FALLBACK_MODEL,
    GEMINI_MAX_RETRIES,
    GEMINI_RETRY_BACKOFF_S,
    GEMINI_TIMEOUT_S,
    IG_FUZZY_MIN_CORE_LEN,
    MAX_IMAGE_SIDE,
    staged_export_column_keys,
    staged_export_new_workbook_column_pairs,
)
from logging_config import get_logger, log_fields

logger = get_logger("processor")

_JSON_FENCE = re.compile(r"```(?:json)?\s*([\s\S]*?)\s*```", re.IGNORECASE)

STAGING_META_KEYS = frozenset({"batch_duplicate"})


def row_without_staging_meta(row: dict[str, Any]) -> dict[str, Any]:
    return {k: v for k, v in row.items() if k not in STAGING_META_KEYS}


def is_staged_row_saveable(row: dict[str, Any]) -> bool:
    """Rows offered for download / merge (exclude in-batch duplicate copies only)."""
    return not row.get("batch_duplicate")


def pending_export_rows(
    new_rows: list[Any],
    duplicate_rows: list[Any],
    scope: str,
    *,
    workbook_sources: list[tuple[str, bytes]] | None = None,
) -> list[dict[str, Any]]:
    """
    Build the row list for download or merge from a pending process result.

    scope: "new" (saveable leads not matching workbooks), "duplicates" (process-time
    duplicate_rows plus saveable new_rows that match workbook_sources), "all" (union).

    When workbook_sources is None, live workbook matching is skipped (only process-time
    duplicate_rows are used for the duplicates branch). Pass the same workbook list as
    duplicate detection / sheet-row-matches so "Duplicates only" matches the table overlay.
    """
    s = (scope or "new").strip().lower()
    if s not in {"new", "duplicates", "all"}:
        s = "new"
    dup_clean: list[dict[str, Any]] = []
    for r in duplicate_rows or []:
        if isinstance(r, dict):
            dup_clean.append(row_without_staging_meta(r))

    new_rows_list = list(new_rows or [])
    n_new = len(new_rows_list)
    live_match: list[dict[str, Any] | None] = [None] * n_new
    if workbook_sources:
        mrows: list[dict[str, Any]] = []
        for r in new_rows_list:
            if isinstance(r, dict):
                mrows.append(
                    {
                        "Instagram": r.get("Instagram", ""),
                        "Mobile": r.get("Mobile", ""),
                        "Email": r.get("Email", ""),
                    }
                )
            else:
                mrows.append({"Instagram": "", "Mobile": "", "Email": ""})
        res = match_lead_rows_to_combined_workbooks(mrows, workbook_sources)
        if not res.get("ok"):
            raise ValueError(str(res.get("error", "Workbook match failed.")))
        matches = res.get("matches")
        if not isinstance(matches, list) or len(matches) != n_new:
            raise ValueError("Workbook match payload was invalid.")
        for i, mi in enumerate(matches):
            if mi is None or not isinstance(mi, dict):
                continue
            raw_er = mi.get("excel_row")
            try:
                er = int(raw_er) if raw_er is not None else 0
            except (TypeError, ValueError):
                continue
            live_match[i] = {
                "source_file": str(mi.get("source_file", "") or ""),
                "excel_row": er,
            }

    new_not_sheet_dup: list[dict[str, Any]] = []
    new_sheet_dup: list[dict[str, Any]] = []
    for i, r in enumerate(new_rows_list):
        if not isinstance(r, dict) or not is_staged_row_saveable(r):
            continue
        base = row_without_staging_meta(r)
        m = live_match[i] if i < len(live_match) else None
        if m:
            enriched = {
                **base,
                "Duplicate Source File": m["source_file"],
                "Duplicate Source Row": m["excel_row"],
            }
            new_sheet_dup.append(enriched)
        else:
            new_not_sheet_dup.append(base)

    if s == "duplicates":
        return dup_clean + new_sheet_dup
    if s == "all":
        return new_not_sheet_dup + dup_clean + new_sheet_dup
    return new_not_sheet_dup


def _configure_genai() -> None:
    from workbook_storage import normalize_api_key

    key = normalize_api_key(os.getenv("GEMINI_API_KEY"))
    if not key:
        raise RuntimeError("GEMINI_API_KEY is not set. Add it to your .env file.")
    # Keep both env names aligned for SDK internals.
    os.environ["GOOGLE_API_KEY"] = key
    genai.configure(api_key=key)


def _list_model_candidates(requested_model: str | None = None) -> list[str]:
    """Primary model + optional single fallback. Skips slow API discovery by default."""
    ordered: list[str] = []
    seen: set[str] = set()

    def push(name: str | None) -> None:
        if not name:
            return
        model_name = name.strip()
        if not model_name or model_name in seen:
            return
        seen.add(model_name)
        ordered.append(model_name)

    push(requested_model or DEFAULT_GEMINI_MODEL)
    push(DEFAULT_GEMINI_MODEL)
    push(GEMINI_FALLBACK_MODEL)

    if GEMINI_DISCOVER_MODELS:
        try:
            discovered: list[str] = []
            for model in genai.list_models():
                name = str(getattr(model, "name", "") or "")
                methods = list(getattr(model, "supported_generation_methods", []) or [])
                if "generateContent" not in methods:
                    continue
                if "gemini" not in name.lower():
                    continue
                discovered.append(name)
            discovered.sort()
            for name in discovered:
                push(name)
        except Exception:
            pass

    if not ordered:
        ordered.append(DEFAULT_GEMINI_MODEL or "models/gemini-2.5-flash")
    return ordered


def normalize_username(value: str | None) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    s = str(value).strip()
    if s.startswith("@"):
        s = s[1:]
    return s.strip().lower()


_IG_TRAILING_DIGITS = re.compile(r"\d+$")


def instagram_core_key(norm: str) -> str:
    """Normalized handle with trailing ASCII digits removed."""
    if not norm:
        return ""
    return _IG_TRAILING_DIGITS.sub("", norm)


def instagram_handles_match_for_duplicate(a: str, b: str) -> bool:
    """
    True when two normalized Instagram handles should count as the same lead.

    Matches exact handles, numbered variants (handle2 vs handle), and near-miss
    spellings often seen in master sheets (e.g. zsgprojects vs zsgproject3).
    """
    if not a or not b:
        return False
    if a == b:
        return True

    short, long = (a, b) if len(a) <= len(b) else (b, a)
    if long.startswith(short) and short:
        suffix = long[len(short) :]
        if suffix.isdigit():
            return True

    ca = instagram_core_key(a)
    cb = instagram_core_key(b)
    if not ca or not cb:
        return False

    min_len = IG_FUZZY_MIN_CORE_LEN
    if ca == cb and a != b:
        if len(ca) >= min_len and (_IG_TRAILING_DIGITS.search(a) or _IG_TRAILING_DIGITS.search(b)):
            return True

    short_c, long_c = (ca, cb) if len(ca) <= len(cb) else (cb, ca)
    if len(short_c) < min_len or not long_c.startswith(short_c):
        return False
    suffix_c = long_c[len(short_c) :]
    if not suffix_c:
        return False
    if suffix_c.isdigit():
        return True
    if len(suffix_c) == 1 and suffix_c.isalnum():
        return True
    return False


def normalize_phone(value: str | None) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    digits = re.sub(r"\D", "", str(value))
    return digits


def phone_match_key(digits: str) -> str:
    """
    9-digit comparison key for duplicate matching.

    10-digit numbers with a leading 0 compare without that digit, so 0123456789
    matches 123456789.
    """
    if not digits:
        return ""
    if len(digits) == 11 and digits.startswith("61") and digits[2] == "4":
        return digits[2:]
    if len(digits) == 10 and digits.startswith("0"):
        return digits[1:]
    if len(digits) == 9:
        return digits
    return digits


def phones_match_for_duplicate(a: str, b: str) -> bool:
    ka = phone_match_key(a)
    kb = phone_match_key(b)
    return bool(ka) and ka == kb


def format_phone_for_storage(raw: str | None) -> str:
    """
    Canonical phone for storage/display (AU-oriented grouping).
    - 10 digits with leading 0: 0123-456-789 (4-3-3)
    - 9 digits: 1234-567-89 (4-3-2)
    """
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return ""
    digits = re.sub(r"\D", "", str(raw))
    if not digits:
        return ""
    if len(digits) == 11 and digits.startswith("61") and digits[2] == "4":
        digits = "0" + digits[2:]
    if len(digits) == 10 and digits.startswith("0"):
        return f"{digits[0:4]}-{digits[4:7]}-{digits[7:10]}"
    if len(digits) == 9:
        return f"{digits[0:4]}-{digits[4:7]}-{digits[7:9]}"
    if len(digits) == 10:
        return f"{digits[0:4]}-{digits[4:7]}-{digits[7:10]}"
    parts: list[str] = []
    d = digits
    while len(d) > 3:
        parts.insert(0, d[-3:])
        d = d[:-3]
    if d:
        parts.insert(0, d)
    return "-".join(parts)


def normalize_email(value: str | None) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    return str(value).strip().lower()


def _parse_json_response(text: str) -> dict[str, Any]:
    text = (text or "").strip()
    m = _JSON_FENCE.search(text)
    if m:
        text = m.group(1).strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        start = text.find("{")
        end = text.rfind("}")
        if start >= 0 and end > start:
            return json.loads(text[start : end + 1])
        raise


def _detect_columns(df: pd.DataFrame) -> tuple[str | None, str | None, str | None]:
    """Returns (username_col, phone_col, error_message_if_any)."""
    if df.empty or len(df.columns) == 0:
        return None, None, "Master sheet is empty."

    lowered = {str(c).strip().lower(): c for c in df.columns}

    def pick(*candidates: str) -> str | None:
        for name in candidates:
            if name in lowered:
                return lowered[name]
        return None

    username_col = pick(
        "username",
        "user",
        "instagram",
        "handle",
        "ig",
        "insta",
        "instagram username",
    )
    phone_col = pick(
        "phone",
        "phone number",
        "phonenumber",
        "mobile",
        "tel",
        "telephone",
    )

    if not username_col:
        for c in df.columns:
            cl = str(c).lower()
            if "user" in cl or "insta" in cl or "handle" in cl:
                username_col = c
                break

    if not phone_col:
        for c in df.columns:
            cl = str(c).lower()
            if "phone" in cl or "mobile" in cl or "tel" in cl:
                phone_col = c
                break

    if not username_col or not phone_col:
        return None, None, (
            "Could not detect Username and Phone columns in master file(s). "
            "Name columns to include 'username' and 'phone' (or similar)."
        )

    return username_col, phone_col, None


def load_master_records(master_uploads: list[Any]) -> tuple[list[dict[str, Any]], str | None]:
    """
    Each master record:
    source_file, excel_row (1-based sheet row), username_norm, phone_norm
    """
    records: list[dict[str, Any]] = []
    username_col_global: str | None = None
    phone_col_global: str | None = None

    for upl in master_uploads:
        name = getattr(upl, "name", "master.xlsx")
        try:
            upl.seek(0)
        except Exception:  # noqa: BLE001
            pass
        raw = upl.read()
        buf = io.BytesIO(raw)
        try:
            df = pd.read_excel(buf, sheet_name=0, dtype=object)
        except Exception as e:  # noqa: BLE001
            return [], f"Failed to read {name}: {e}"

        if username_col_global is None:
            u_col, p_col, err = _detect_columns(df)
            if err:
                return [], err
            username_col_global, phone_col_global = u_col, p_col
        else:
            u_col, p_col = username_col_global, phone_col_global
            if u_col not in df.columns or p_col not in df.columns:
                return [], (
                    f"{name} is missing expected columns "
                    f"'{username_col_global}' / '{phone_col_global}'."
                )

        for i, row in df.iterrows():
            excel_row = int(i) + 2  # header on row 1
            u = normalize_username(row.get(u_col))
            p = normalize_phone(row.get(p_col))
            records.append(
                {
                    "source_file": name,
                    "excel_row": excel_row,
                    "username_norm": u,
                    "phone_norm": p,
                }
            )

    return records, None


def build_master_index_from_dataframe(
    df: pd.DataFrame,
    source_file: str,
) -> tuple[list[dict[str, Any]], str | None, str | None, str | None]:
    """
    Returns:
        records, username_col, phone_col, error
    """
    username_col, phone_col, err = _detect_columns(df)
    if err:
        return [], None, None, err

    records: list[dict[str, Any]] = []
    for i, row in df.iterrows():
        records.append(
            {
                "source_file": source_file,
                "excel_row": int(i) + 2,
                "username_norm": normalize_username(row.get(username_col)),
                "phone_norm": normalize_phone(row.get(phone_col)),
            }
        )
    return records, username_col, phone_col, None


def find_duplicate(
    username_norm: str,
    phone_norm: str,
    master_records: list[dict[str, Any]],
) -> dict[str, Any] | None:
    for rec in master_records:
        rec_user = rec.get("username_norm") or ""
        if username_norm and rec_user and instagram_handles_match_for_duplicate(
            username_norm, rec_user
        ):
            return {"source_file": rec["source_file"], "excel_row": rec["excel_row"]}
        if phone_norm and phones_match_for_duplicate(phone_norm, rec.get("phone_norm", "")):
            return {"source_file": rec["source_file"], "excel_row": rec["excel_row"]}
    return None


def _gemini_request_options() -> dict[str, int]:
    return {"timeout": GEMINI_TIMEOUT_S}


def _call_gemini_generate(
    model: Any,
    prompt: str,
    img: Image.Image,
    gen_cfg: Any,
) -> Any:
    """Single Gemini vision call with a hard timeout (no silent double-call)."""
    return model.generate_content(
        [prompt, img],
        generation_config=gen_cfg,
        request_options=_gemini_request_options(),
    )


def _is_transient_gemini_error(exc: BaseException) -> bool:
    msg = str(exc).lower()
    needles = (
        "429",
        "500",
        "502",
        "503",
        "504",
        "timeout",
        "timed out",
        "resource exhausted",
        "rate limit",
        "too many requests",
        "unavailable",
        "deadline",
    )
    return any(n in msg for n in needles)


def extract_from_screenshot(
    image_bytes: bytes,
    filename: str,
    model_name: str,
    model_candidates: list[str] | None = None,
) -> dict[str, str]:
    _configure_genai()
    try:
        img = Image.open(io.BytesIO(image_bytes))
        if img.mode not in ("RGB", "L"):
            img = img.convert("RGB")
        if MAX_IMAGE_SIDE > 0:
            max_side = max(img.size)
            if max_side > MAX_IMAGE_SIDE:
                img.thumbnail((MAX_IMAGE_SIDE, MAX_IMAGE_SIDE))
        prompt = (
            f"{EXTRACTION_SYSTEM_INSTRUCTION}\n\n"
            f"Image filename (context only): {filename}\n"
            "Extract business_name, mobile, email, instagram from this screenshot."
        )
        gen_cfg = genai.GenerationConfig(
            temperature=0.1,
            response_mime_type="application/json",
        )
        candidates = model_candidates or _list_model_candidates(model_name)
        if not candidates:
            raise RuntimeError("No Gemini model candidate available. Set GEMINI_MODEL in .env.")

        response = None
        used_model = ""
        errors: list[str] = []
        for candidate in candidates:
            model = genai.GenerativeModel(model_name=candidate)
            last_err: Exception | None = None
            for attempt in range(1, GEMINI_MAX_RETRIES + 1):
                try:
                    response = _call_gemini_generate(model, prompt, img, gen_cfg)
                    used_model = candidate
                    last_err = None
                    break
                except Exception as candidate_err:  # noqa: BLE001
                    last_err = candidate_err
                    logger.warning(
                        "gemini_candidate_attempt_failed %s",
                        log_fields(
                            image=filename,
                            model=candidate,
                            attempt=attempt,
                            max_retries=GEMINI_MAX_RETRIES,
                            timeout_s=GEMINI_TIMEOUT_S,
                            error=str(candidate_err),
                        ),
                    )
                    if attempt < GEMINI_MAX_RETRIES and _is_transient_gemini_error(candidate_err):
                        time.sleep(GEMINI_RETRY_BACKOFF_S * attempt)
                        continue
                    break
            if response is not None and used_model:
                break
            if last_err is not None:
                errors.append(f"{candidate}: {last_err}")

        if response is None:
            detail = " | ".join(errors) if errors else "Gemini call failed."
            logger.warning(
                "gemini_all_models_failed %s",
                log_fields(image=filename, candidates_tried=len(candidates), error=detail),
            )
            raise RuntimeError(detail)
        text = getattr(response, "text", None) or ""
        data = _parse_json_response(text)
        business_name = str(data.get("business_name", "") or "").strip()
        mobile_raw = str(data.get("mobile", data.get("phone", "")) or "").strip()
        mobile = format_phone_for_storage(mobile_raw)
        email = str(data.get("email", "") or "").strip()
        instagram = str(data.get("instagram", data.get("username", "")) or "").strip()
        notes = str(data.get("notes", "") or "").strip()
        err_parts = []
        if notes:
            err_parts.append(notes)
        if getattr(response, "prompt_feedback", None):
            pf = response.prompt_feedback
            if getattr(pf, "block_reason", None):
                err_parts.append(f"blocked:{pf.block_reason}")
        return {
            "business_name": business_name,
            "mobile": mobile,
            "email": email,
            "instagram": instagram.lstrip("@"),
            "error": "; ".join(err_parts) if err_parts else "",
            "_model_used": used_model,
        }
    except Exception as e:  # noqa: BLE001
        logger.warning(
            "screenshot_extraction_exception %s",
            log_fields(image=filename, error=str(e)),
        )
        return {
            "business_name": "",
            "mobile": "",
            "email": "",
            "instagram": "",
            "error": str(e),
            "_model_used": "",
        }


def _dataframe_to_xlsx_bytes(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")
    buf.seek(0)
    return buf.read()


def _append_new_rows_to_workbook(
    workbook_bytes: bytes,
    new_rows: list[dict[str, Any]],
    username_col: str,
    phone_col: str,
) -> bytes:
    """
    Appends only new rows into first sheet of uploaded workbook.
    Writes required columns (username/phone) and fills optional columns when present.
    """
    if not new_rows:
        return workbook_bytes

    wb = load_workbook(filename=io.BytesIO(workbook_bytes))
    ws = wb.worksheets[0]
    header_map: dict[str, int] = {}
    for idx, cell in enumerate(ws[1], start=1):
        header_map[str(cell.value).strip().lower() if cell.value is not None else ""] = idx

    uname_idx = header_map.get(str(username_col).strip().lower())
    phone_idx = header_map.get(str(phone_col).strip().lower())
    if not uname_idx or not phone_idx:
        raise RuntimeError("Could not find username/phone columns while appending to workbook.")

    optional_field_to_headers = {
        "Image Name": ["image name", "image", "screenshot", "file name", "filename"],
        "Error": ["error", "notes", "note"],
        "Date & Time": ["date & time", "datetime", "date time", "timestamp", "date"],
    }
    optional_indexes: dict[str, int] = {}
    for field, header_candidates in optional_field_to_headers.items():
        for h in header_candidates:
            if h in header_map:
                optional_indexes[field] = header_map[h]
                break

    for row in new_rows:
        ws.append([""] * ws.max_column)
        r = ws.max_row
        ws.cell(row=r, column=uname_idx, value=row.get("Username", ""))
        ws.cell(row=r, column=phone_idx, value=row.get("Phone Number", ""))
        for field, col_idx in optional_indexes.items():
            ws.cell(row=r, column=col_idx, value=row.get(field, ""))

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


def process_batch(
    master_uploads: list[Any],
    screenshot_uploads: list[Any],
    model_name: str | None = None,
) -> dict[str, Any]:
    model = (model_name or DEFAULT_GEMINI_MODEL or "").strip()
    _configure_genai()
    candidates = _list_model_candidates(model)
    master_records, master_err = load_master_records(master_uploads)
    if master_err:
        return {"ok": False, "error": master_err}

    now = datetime.now(timezone.utc).astimezone()
    ts = now.strftime("%Y-%m-%d %H:%M:%S")

    new_rows: list[dict[str, Any]] = []
    dup_rows: list[dict[str, Any]] = []
    failed_rows: list[dict[str, Any]] = []
    failed_rows: list[dict[str, Any]] = []

    for upl in screenshot_uploads:
        fname = getattr(upl, "name", "image.png")
        try:
            upl.seek(0)
        except Exception:  # noqa: BLE001
            pass
        raw = upl.read()
        extracted = extract_from_screenshot(raw, fname, model, model_candidates=candidates)
        used = extracted.get("_model_used", "")
        if used and used in candidates and candidates[0] != used:
            candidates.remove(used)
            candidates.insert(0, used)
        u_norm = normalize_username(extracted["username"])
        p_norm = normalize_phone(extracted["phone"])
        dup = find_duplicate(u_norm, p_norm, master_records)

        base = {
            "Image Name": fname,
            "Username": extracted["username"],
            "Phone Number": extracted["phone"],
            "Error": extracted["error"],
            "Date & Time": ts,
        }

        if dup:
            dup_rows.append(
                {
                    **base,
                    "Duplicate Source File": dup["source_file"],
                    "Duplicate Source Row": dup["excel_row"],
                }
            )
        else:
            new_rows.append(base)

    new_df = pd.DataFrame(new_rows)
    dup_df = pd.DataFrame(dup_rows)

    # Stable column order for duplicates sheet (matches client template).
    dup_cols = [
        "Image Name",
        "Username",
        "Phone Number",
        "Error",
        "Duplicate Source File",
        "Duplicate Source Row",
        "Date & Time",
    ]
    for col in dup_cols:
        if col not in dup_df.columns:
            dup_df[col] = ""
    dup_df = dup_df[[c for c in dup_cols if c in dup_df.columns]]

    new_cols = ["Image Name", "Username", "Phone Number", "Error", "Date & Time"]
    for col in new_cols:
        if col not in new_df.columns:
            new_df[col] = ""
    new_df = new_df[[c for c in new_cols if c in new_df.columns]]

    return {
        "ok": True,
        "new_bytes": _dataframe_to_xlsx_bytes(new_df),
        "dup_bytes": _dataframe_to_xlsx_bytes(dup_df),
        "counts": {"new": len(new_df), "duplicates": len(dup_df)},
    }


def process_single_master_and_screenshots(
    master_bytes: bytes,
    master_filename: str,
    screenshot_items: list[tuple[str, bytes]],
    model_name: str | None = None,
) -> dict[str, Any]:
    model = (model_name or DEFAULT_GEMINI_MODEL or "").strip()
    _configure_genai()
    candidates = _list_model_candidates(model)
    try:
        master_df = pd.read_excel(io.BytesIO(master_bytes), sheet_name=0, dtype=object)
    except Exception as e:  # noqa: BLE001
        return {"ok": False, "error": f"Failed to read master workbook: {e}"}

    master_records, username_col, phone_col, index_err = build_master_index_from_dataframe(
        master_df,
        source_file=master_filename,
    )
    if index_err:
        return {"ok": False, "error": index_err}

    now = datetime.now(timezone.utc).astimezone()
    ts = now.strftime("%Y-%m-%d %H:%M:%S")
    new_rows: list[dict[str, Any]] = []
    dup_rows: list[dict[str, Any]] = []
    failed_rows: list[dict[str, Any]] = []

    for filename, raw in screenshot_items:
        extracted = extract_from_screenshot(raw, filename, model, model_candidates=candidates)
        used = extracted.get("_model_used", "")
        if used and used in candidates and candidates[0] != used:
            candidates.remove(used)
            candidates.insert(0, used)
        u_norm = normalize_username(extracted["username"])
        p_norm = normalize_phone(extracted["phone"])
        dup = find_duplicate(u_norm, p_norm, master_records)

        base = {
            "Image Name": filename,
            "Username": extracted["username"],
            "Phone Number": extracted["phone"],
            "Error": extracted["error"],
            "Date & Time": ts,
        }

        if dup:
            dup_rows.append(
                {
                    **base,
                    "Duplicate Source File": dup["source_file"],
                    "Duplicate Source Row": dup["excel_row"],
                }
            )
        else:
            new_rows.append(base)
            # Keep in-memory index updated to avoid duplicates within same batch.
            master_records.append(
                {
                    "source_file": master_filename,
                    "excel_row": len(master_records) + 2,
                    "username_norm": u_norm,
                    "phone_norm": p_norm,
                }
            )

    new_df = pd.DataFrame(new_rows)
    dup_df = pd.DataFrame(dup_rows)
    updated_master_bytes = _append_new_rows_to_workbook(
        workbook_bytes=master_bytes,
        new_rows=new_rows,
        username_col=username_col or "",
        phone_col=phone_col or "",
    )

    return {
        "ok": True,
        "counts": {"new": len(new_rows), "duplicates": len(dup_rows), "processed": len(screenshot_items)},
        "new_rows": new_rows,
        "duplicate_rows": dup_rows,
        "new_bytes": _dataframe_to_xlsx_bytes(new_df),
        "dup_bytes": _dataframe_to_xlsx_bytes(dup_df),
        "updated_master_bytes": updated_master_bytes,
    }


def _detect_required_columns(
    df: pd.DataFrame,
) -> tuple[dict[str, str | None] | None, str | None]:
    """
    Map sheet columns to logical fields.

    For duplicate checking against a workbook, only Instagram (or equivalent) and
    Mobile/Phone are required. Email and Business Name are optional (treated as empty
    when absent). Appending new rows still requires all four — see _append_required_rows_to_workbook.
    """
    if len(df.columns) == 0:
        return None, "Master sheet has no headers."

    lowered = {str(c).strip().lower(): str(c) for c in df.columns}

    def pick(candidates: list[str]) -> str | None:
        for name in candidates:
            if name in lowered:
                return lowered[name]
        for col in df.columns:
            col_l = str(col).strip().lower()
            for name in candidates:
                if len(name) < 6 and name != col_l:
                    continue
                if name in col_l:
                    return str(col)
        return None

    mapping: dict[str, str | None] = {
        "business_name": pick(["business name", "business", "company name", "company"]),
        "mobile": pick(
            [
                "phone numbers",
                "mobile",
                "phone",
                "phone number",
                "phonenumber",
                "telephone",
                "tel",
            ]
        ),
        "email": pick(["email", "e-mail", "mail"]),
        "instagram": pick(
            [
                "instagram usernames",
                "instagram",
                "insta",
                "username",
                "handle",
                "ig",
            ]
        ),
    }
    missing_required = [k for k in ("instagram", "mobile") if not mapping.get(k)]
    if missing_required:
        return None, (
            "Could not detect required columns: "
            + ", ".join(missing_required)
            + ". Each sheet needs Instagram (or Instagram Usernames, handle, etc.) "
            "and Mobile or Phone Numbers. Email and Business Name are optional. "
            "Duplicate vs new: same Instagram handle (including numbered variants such as "
            "handle2 vs handle), or same phone (9-digit match, ignoring a leading 0 on "
            "10-digit numbers)."
        )
    return mapping, None


def _build_master_index_required(
    df: pd.DataFrame,
    source_file: str,
    colmap: dict[str, str | None],
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    ig_col = colmap.get("instagram")
    mob_col = colmap.get("mobile")
    em_col = colmap.get("email")
    for i, row in df.iterrows():
        ig_raw = row.get(ig_col, "") if ig_col else ""
        mob_raw = row.get(mob_col, "") if mob_col else ""
        em_raw = row.get(em_col, "") if em_col else ""
        records.append(
            {
                "source_file": source_file,
                "excel_row": int(i) + 2,
                "instagram_norm": normalize_username(ig_raw),
                "mobile_norm": normalize_phone(mob_raw),
                "email_norm": normalize_email(em_raw),
            }
        )
    return records


def _find_duplicate_required(
    instagram_norm: str,
    mobile_norm: str,
    email_norm: str,
    records: list[dict[str, Any]],
) -> dict[str, Any] | None:
    """
    Duplicate vs new: same Instagram handle (exact or fuzzy numbered/near-miss variant),
    or same phone (9-digit key; drop leading 0 on 10-digit).
    """
    del email_norm
    for rec in records:
        rec_ig = rec.get("instagram_norm") or rec.get("username_norm") or ""
        if instagram_norm and rec_ig and instagram_handles_match_for_duplicate(
            instagram_norm, rec_ig
        ):
            return {"source_file": rec["source_file"], "excel_row": rec["excel_row"]}
        rec_mob = rec.get("mobile_norm") or rec.get("phone_norm") or ""
        if mobile_norm and phones_match_for_duplicate(mobile_norm, rec_mob):
            return {"source_file": rec["source_file"], "excel_row": rec["excel_row"]}
    return None


def build_master_records_from_workbook_sources(
    sources: list[tuple[str, bytes]],
) -> dict[str, Any]:
    """
    Merge first-sheet lead indexes from multiple workbooks.
    Each tuple is (display filename, .xlsx bytes).
    """
    master_records: list[dict[str, Any]] = []
    for display_name, wb_bytes in sources:
        label = (display_name or "").strip() or "workbook"
        try:
            master_df = pd.read_excel(io.BytesIO(wb_bytes), sheet_name=0, dtype=object)
        except Exception as e:  # noqa: BLE001
            return {"ok": False, "error": f"{label}: could not read workbook ({e})"}
        colmap, col_err = _detect_required_columns(master_df)
        if col_err or not colmap:
            return {"ok": False, "error": f"{label}: {col_err or 'column detection failed'}"}
        master_records.extend(_build_master_index_required(master_df, label, colmap))
    return {"ok": True, "records": master_records}


def match_lead_rows_to_combined_workbooks(
    rows: list[dict[str, Any]],
    workbook_sources: list[tuple[str, bytes]],
) -> dict[str, Any]:
    """Match each row against the union of all workbook_sources (first sheet each)."""
    if not workbook_sources:
        return {"ok": True, "matches": [None] * len(rows)}
    idx = build_master_records_from_workbook_sources(workbook_sources)
    if not idx.get("ok"):
        return {"ok": False, "error": idx.get("error", "Failed to build workbook index.")}
    records: list[dict[str, Any]] = idx["records"]
    matches: list[dict[str, Any] | None] = []
    for row in rows:
        ig = normalize_username(row.get("Instagram"))
        mob = normalize_phone(row.get("Mobile"))
        em = normalize_email(row.get("Email"))
        dup = _find_duplicate_required(ig, mob, em, records)
        if dup:
            matches.append(
                {
                    "excel_row": int(dup["excel_row"]),
                    "source_file": str(dup["source_file"]),
                }
            )
        else:
            matches.append(None)
    return {"ok": True, "matches": matches}


def match_lead_rows_to_workbook(
    rows: list[dict[str, Any]],
    workbook_bytes: bytes,
    workbook_display_name: str,
) -> dict[str, Any]:
    """
    For each row dict (keys Instagram, Mobile, Email), find the first matching row
    in the workbook's first sheet using the same rules as process-time deduplication.
    """
    return match_lead_rows_to_combined_workbooks(rows, [(workbook_display_name, workbook_bytes)])


def _append_required_rows_to_workbook(
    workbook_bytes: bytes,
    new_rows: list[dict[str, Any]],
    colmap: dict[str, str | None],
) -> bytes:
    if not new_rows:
        return workbook_bytes
    for key, label in (
        ("business_name", "Business Name"),
        ("mobile", "Mobile / Phone"),
        ("email", "Email"),
        ("instagram", "Instagram"),
    ):
        if not colmap.get(key):
            raise RuntimeError(
                f"Cannot append rows: this workbook is missing a {label} column (or an equivalent "
                "header we recognize). Add the column or choose a different target workbook."
            )
    wb = load_workbook(filename=io.BytesIO(workbook_bytes))
    ws = wb.worksheets[0]
    header_map: dict[str, int] = {}
    for idx, cell in enumerate(ws[1], start=1):
        key = str(cell.value).strip().lower() if cell.value is not None else ""
        header_map[key] = idx

    b_key = str(colmap["business_name"]).strip().lower()
    m_key = str(colmap["mobile"]).strip().lower()
    e_key = str(colmap["email"]).strip().lower()
    i_key = str(colmap["instagram"]).strip().lower()
    b_idx = header_map.get(b_key)
    m_idx = header_map.get(m_key)
    e_idx = header_map.get(e_key)
    i_idx = header_map.get(i_key)
    if not all([b_idx, m_idx, e_idx, i_idx]):
        raise RuntimeError("Could not map required columns while appending rows.")

    for row in new_rows:
        ws.append([""] * ws.max_column)
        r = ws.max_row
        ws.cell(row=r, column=b_idx, value=row.get("Business Name", ""))
        ws.cell(row=r, column=m_idx, value=row.get("Mobile", ""))
        ws.cell(row=r, column=e_idx, value=row.get("Email", ""))
        ws.cell(row=r, column=i_idx, value=row.get("Instagram", ""))

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


def process_single_master_and_screenshots(
    master_bytes: bytes,
    master_filename: str,
    screenshot_items: list[tuple[str, bytes]],
    model_name: str | None = None,
) -> dict[str, Any]:
    model = (model_name or DEFAULT_GEMINI_MODEL or "").strip()
    try:
        master_df = pd.read_excel(io.BytesIO(master_bytes), sheet_name=0, dtype=object)
    except Exception as e:  # noqa: BLE001
        return {"ok": False, "error": f"Failed to read master workbook: {e}"}

    colmap, col_err = _detect_required_columns(master_df)
    if col_err or not colmap:
        return {"ok": False, "error": col_err or "Column detection failed."}

    master_records = _build_master_index_required(master_df, master_filename, colmap)
    now = datetime.now(timezone.utc).astimezone()
    ts = now.strftime("%Y-%m-%d %H:%M:%S")

    new_rows: list[dict[str, Any]] = []
    dup_rows: list[dict[str, Any]] = []
    failed_rows: list[dict[str, Any]] = []

    for filename, raw in screenshot_items:
        extracted = extract_from_screenshot(raw, filename, model)
        ig_norm = normalize_username(extracted["instagram"])
        mob_norm = normalize_phone(extracted["mobile"])
        em_norm = normalize_email(extracted["email"])
        dup = _find_duplicate_required(ig_norm, mob_norm, em_norm, master_records)

        base = {
            "Image Name": filename,
            "Business Name": extracted["business_name"],
            "Mobile": extracted["mobile"],
            "Email": extracted["email"],
            "Instagram": extracted["instagram"],
            "Error": extracted["error"],
            "Date & Time": ts,
        }

        has_data = bool(
            str(extracted.get("business_name", "")).strip()
            or str(extracted.get("mobile", "")).strip()
            or str(extracted.get("email", "")).strip()
            or str(extracted.get("instagram", "")).strip()
        )
        if not has_data and str(extracted.get("error", "")).strip():
            failed_rows.append(base)
            continue

        if dup:
            dup_rows.append(
                {
                    **base,
                    "Duplicate Source File": dup["source_file"],
                    "Duplicate Source Row": dup["excel_row"],
                }
            )
        else:
            new_rows.append(base)
            master_records.append(
                {
                    "source_file": master_filename,
                    "excel_row": len(master_records) + 2,
                    "instagram_norm": ig_norm,
                    "mobile_norm": mob_norm,
                    "email_norm": em_norm,
                }
            )

    new_df = pd.DataFrame(new_rows)
    dup_df = pd.DataFrame(dup_rows)
    updated_master_bytes = _append_required_rows_to_workbook(master_bytes, new_rows, colmap)

    return {
        "ok": True,
        "counts": {
            "new": len(new_rows),
            "duplicates": len(dup_rows),
            "failed": len(failed_rows),
            "processed": len(screenshot_items),
        },
        "new_rows": new_rows,
        "duplicate_rows": dup_rows,
        "failed_rows": failed_rows,
        "new_bytes": _dataframe_to_xlsx_bytes(new_df),
        "dup_bytes": _dataframe_to_xlsx_bytes(dup_df),
        "updated_master_bytes": updated_master_bytes,
    }


def master_records_with_staged_batch(
    workbook_master_records: list[dict[str, Any]],
    staged_new_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Extend workbook index with saveable leads from an in-progress batch (chunk continuation)."""
    records = list(workbook_master_records)
    for row in staged_new_rows:
        if not isinstance(row, dict) or row.get("batch_duplicate"):
            continue
        ig_norm = normalize_username(str(row.get("Instagram", "")))
        mob_norm = normalize_phone(str(row.get("Mobile", "")))
        if not ig_norm and not mob_norm:
            continue
        records.append(
            {
                "source_file": BATCH_DEDUPE_SOURCE_LABEL,
                "excel_row": len(records) + 2,
                "instagram_norm": ig_norm,
                "mobile_norm": mob_norm,
                "email_norm": normalize_email(str(row.get("Email", ""))),
            }
        )
    return records


def _row_has_lead_data(extracted: dict[str, Any]) -> bool:
    return bool(
        str(extracted.get("business_name", "")).strip()
        or str(extracted.get("mobile", "")).strip()
        or str(extracted.get("email", "")).strip()
        or str(extracted.get("instagram", "")).strip()
    )


def _classify_extractions(
    extractions: list[tuple[int, str, dict[str, Any], int, str]],
    master_records: list[dict[str, Any]],
    ts: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    """Pass 2: sequential dedup in original upload order."""
    new_rows: list[dict[str, Any]] = []
    dup_rows: list[dict[str, Any]] = []
    failed_rows: list[dict[str, Any]] = []
    total = len(extractions)

    for index, filename, extracted, duration_ms, used in extractions:
        source_label = "gemini"
        base = {
            "Image Name": filename,
            "Business Name": extracted["business_name"],
            "Mobile": extracted["mobile"],
            "Email": extracted["email"],
            "Instagram": extracted["instagram"],
            "Error": extracted["error"],
            "Date & Time": ts,
        }

        if not _row_has_lead_data(extracted) and str(extracted.get("error", "")).strip():
            failed_rows.append(base)
            logger.warning(
                "screenshot_failed %s",
                log_fields(
                    index=index,
                    total=total,
                    image=filename,
                    source=source_label,
                    model=used,
                    error=extracted.get("error", ""),
                    duration_ms=duration_ms,
                ),
            )
            continue

        ig_norm = normalize_username(extracted["instagram"])
        mob_norm = normalize_phone(extracted["mobile"])
        em_norm = normalize_email(extracted["email"])
        dup = _find_duplicate_required(ig_norm, mob_norm, em_norm, master_records)

        if dup:
            if dup["source_file"] != BATCH_DEDUPE_SOURCE_LABEL:
                dup_rows.append(
                    {
                        **base,
                        "Duplicate Source File": dup["source_file"],
                        "Duplicate Source Row": dup["excel_row"],
                    }
                )
                logger.info(
                    "screenshot_completed %s",
                    log_fields(
                        index=index,
                        total=total,
                        image=filename,
                        source=source_label,
                        model=used,
                        outcome="duplicate",
                        duplicate_file=dup["source_file"],
                        duplicate_row=dup["excel_row"],
                        duration_ms=duration_ms,
                    ),
                )
            else:
                new_rows.append({**base, "batch_duplicate": True})
                logger.info(
                    "screenshot_completed %s",
                    log_fields(
                        index=index,
                        total=total,
                        image=filename,
                        outcome="in_batch_duplicate",
                        duration_ms=duration_ms,
                    ),
                )
            continue

        new_rows.append(base)
        logger.info(
            "screenshot_completed %s",
            log_fields(
                index=index,
                total=total,
                image=filename,
                source=source_label,
                model=used,
                outcome="new",
                instagram=extracted.get("instagram", ""),
                duration_ms=duration_ms,
            ),
        )
        master_records.append(
            {
                "source_file": BATCH_DEDUPE_SOURCE_LABEL,
                "excel_row": len(master_records) + 2,
                "instagram_norm": ig_norm,
                "mobile_norm": mob_norm,
                "email_norm": em_norm,
            }
        )

    return new_rows, dup_rows, failed_rows


async def _extract_screenshots_parallel(
    screenshot_items: list[tuple[str, bytes]],
    model: str,
    candidates: list[str],
    concurrency: int,
) -> list[tuple[int, str, dict[str, Any], int, str]]:
    """Pass 1: bounded-concurrency Gemini extraction (order preserved)."""
    total = len(screenshot_items)
    sem = asyncio.Semaphore(concurrency)
    candidates_lock = threading.Lock()

    async def run_one(index: int, filename: str, raw: bytes) -> tuple[int, str, dict[str, Any], int, str]:
        image_started = time.perf_counter()
        async with sem:
            logger.info(
                "screenshot_gemini_started %s",
                log_fields(index=index, total=total, image=filename, bytes=len(raw)),
            )
            with candidates_lock:
                cands = list(candidates)
            extracted = await asyncio.to_thread(
                extract_from_screenshot,
                raw,
                filename,
                model,
                cands,
            )
            used = str(extracted.get("_model_used", "") or "")
            if used:
                with candidates_lock:
                    if used in candidates and candidates[0] != used:
                        candidates.remove(used)
                        candidates.insert(0, used)
        duration_ms = int((time.perf_counter() - image_started) * 1000)
        return index, filename, extracted, duration_ms, used

    tasks = [
        run_one(i, filename, raw)
        for i, (filename, raw) in enumerate(screenshot_items, start=1)
    ]
    results = await asyncio.gather(*tasks)
    return sorted(results, key=lambda item: item[0])


async def process_screenshots_only_async(
    screenshot_items: list[tuple[str, bytes]],
    model_name: str | None = None,
    duplicate_workbook_sources: list[tuple[str, bytes]] | None = None,
    initial_master_records: list[dict[str, Any]] | None = None,
    *,
    concurrency: int | None = None,
) -> dict[str, Any]:
    """
    Vision extraction + duplicate detection within the batch and against one or more
    workbooks' first sheets (combined index).

    Extraction runs in parallel (bounded); dedup runs sequentially in upload order.
    """
    model = (model_name or DEFAULT_GEMINI_MODEL or "").strip()
    parallel = max(1, concurrency if concurrency is not None else GEMINI_CONCURRENCY)
    now = datetime.now(timezone.utc).astimezone()
    ts = now.strftime("%Y-%m-%d %H:%M:%S")
    batch_started = time.perf_counter()
    screenshot_names = [name for name, _ in screenshot_items]

    sources = list(duplicate_workbook_sources or [])
    workbook_names = [name for name, _ in sources]
    logger.info(
        "process_batch_started %s",
        log_fields(
            screenshot_count=len(screenshot_items),
            screenshots=screenshot_names,
            gemini_model=model,
            gemini_concurrency=parallel,
            duplicate_workbook_count=len(workbook_names),
            duplicate_workbooks=workbook_names or ["(all MASTER folder workbooks)"],
        ),
    )

    if initial_master_records is not None:
        master_records = list(initial_master_records)
        logger.info(
            "process_batch_workbook_index_ready %s",
            log_fields(master_record_count=len(master_records), continuation=True),
        )
    else:
        idx = build_master_records_from_workbook_sources(sources)
        if not idx.get("ok"):
            err = str(idx.get("error", "Workbook index failed."))
            logger.error("process_batch_workbook_index_failed %s", log_fields(error=err))
            return {"ok": False, "error": err}
        master_records = list(idx["records"])
        logger.info(
            "process_batch_workbook_index_ready %s",
            log_fields(master_record_count=len(master_records)),
        )

    try:
        _configure_genai()
        candidates = _list_model_candidates(model)
        if not candidates:
            raise RuntimeError("No Gemini model candidate available. Set GEMINI_MODEL in .env.")
    except RuntimeError as e:
        logger.error("process_batch_aborted %s", log_fields(error=str(e)))
        return {"ok": False, "error": str(e)}

    extractions = await _extract_screenshots_parallel(
        screenshot_items, model, candidates, parallel
    )
    new_rows, dup_rows, failed_rows = _classify_extractions(extractions, master_records, ts)

    batch_dup_ct = sum(1 for r in new_rows if r.get("batch_duplicate"))
    genuine_new = sum(1 for r in new_rows if is_staged_row_saveable(r))
    counts = {
        "new": genuine_new,
        "duplicates": len(dup_rows),
        "failed": len(failed_rows),
        "processed": len(screenshot_items),
        "in_batch_duplicates": batch_dup_ct,
    }
    logger.info(
        "process_batch_finished %s",
        log_fields(
            duration_ms=int((time.perf_counter() - batch_started) * 1000),
            gemini_concurrency=parallel,
            **counts,
        ),
    )
    return {
        "ok": True,
        "counts": counts,
        "new_rows": new_rows,
        "duplicate_rows": dup_rows,
        "failed_rows": failed_rows,
        "master_records": master_records,
    }


def process_screenshots_only(
    screenshot_items: list[tuple[str, bytes]],
    model_name: str | None = None,
    duplicate_workbook_sources: list[tuple[str, bytes]] | None = None,
    initial_master_records: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    """Sync wrapper for CLI scripts."""
    return asyncio.run(
        process_screenshots_only_async(
            screenshot_items=screenshot_items,
            model_name=model_name,
            duplicate_workbook_sources=duplicate_workbook_sources,
            initial_master_records=initial_master_records,
        )
    )


def staged_rows_to_new_workbook_bytes(rows: list[dict[str, Any]]) -> bytes:
    pairs = staged_export_new_workbook_column_pairs()
    headers = [h for h, _ in pairs]
    if not rows:
        return _dataframe_to_xlsx_bytes(pd.DataFrame(columns=headers))
    data = [
        {h: row.get(k, "") for h, k in pairs}
        for row in (row_without_staging_meta(r) for r in rows)
    ]
    return _dataframe_to_xlsx_bytes(pd.DataFrame(data))


def merge_staged_rows_into_workbook(
    workbook_bytes: bytes,
    staged_rows: list[dict[str, Any]],
) -> dict[str, Any]:
    """
    Append staged leads to an existing workbook, skipping rows that duplicate the file.
    """
    if not staged_rows:
        return {"ok": True, "updated_master_bytes": workbook_bytes, "appended": 0, "skipped_duplicates": 0}

    try:
        master_df = pd.read_excel(io.BytesIO(workbook_bytes), sheet_name=0, dtype=object)
    except Exception as e:  # noqa: BLE001
        return {"ok": False, "error": f"Failed to read workbook: {e}"}

    colmap, col_err = _detect_required_columns(master_df)
    if col_err or not colmap:
        return {"ok": False, "error": col_err or "Column detection failed."}

    master_records = _build_master_index_required(master_df, "existing", colmap)
    rows_to_append: list[dict[str, Any]] = []
    skipped_duplicates = 0

    for row in staged_rows:
        ig_norm = normalize_username(row.get("Instagram", ""))
        mob_norm = normalize_phone(row.get("Mobile", ""))
        em_norm = normalize_email(row.get("Email", ""))
        if _find_duplicate_required(ig_norm, mob_norm, em_norm, master_records):
            skipped_duplicates += 1
            continue
        rows_to_append.append(row)
        master_records.append(
            {
                "source_file": "existing",
                "excel_row": len(master_records) + 2,
                "instagram_norm": ig_norm,
                "mobile_norm": mob_norm,
                "email_norm": em_norm,
            }
        )

    try:
        updated = _append_required_rows_to_workbook(workbook_bytes, rows_to_append, colmap)
    except Exception as e:  # noqa: BLE001
        return {"ok": False, "error": str(e)}

    return {
        "ok": True,
        "updated_master_bytes": updated,
        "appended": len(rows_to_append),
        "skipped_duplicates": skipped_duplicates,
    }
